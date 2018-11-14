# coding=utf-8
# coding=utf-8
# coding=utf-8

import requests
import  unittest
import  json
from openpyxl import load_workbook
import sys
reload(sys)
sys.setdefaultencoding('utf8')
import csv
reload(sys)
sys.setdefaultencoding('utf8')
class testClass(object):
    #封装http POST 函数，返回请求response
    def httpPost(self,keyword,url):
        # data=json.dumps(keyword)
        headers={"Content-Type":"application/json"}
        res=requests.post(url,data=keyword,headers=headers)
        responseJson=res.json()
        return responseJson

    #封装http Get函数，返回response
    def httpGet(self,url,data):
        headers={"Content-Type":"application/json"}
        res=requests.get(url,data=data,headers=headers)
        print  res
        responseJson=res.json()
        print  responseJson
        return responseJson


    #封装请求函数
    def httpGetOrPost(self,method,url,data):
      #  global mres
        headers = {"Content-Type": "application/json"}
        if method in "get":
            mres=requests.get(url,data=data,headers=headers)
        elif method == "post":
            # postdata = json.dumps(data)
            mres=requests.post(url,data=data,headers=headers)
        elif method in"put":
            mres=requests.put(url,data=data,headers=headers)
        elif method in "delete":
            mres=requests.delete(url,data=data,headers=headers)
        else:
            mres = requests.post(url, data=data, headers=headers)
            print("错误")
       # responseJson=mres.json()
        return mres.json()


    #读excel
    def readSheetdata(self,cell):
        wb=load_workbook(r'd:\apitestcase.xlsx')
        sheet=wb.active
        value=sheet[cell]
        # print(value.value)
        return value.value

    def writeSheetdata(self,cell,data):
        wb=load_workbook()
        sheet=wb.active
        sheet[cell].value=data
        wb.save


    #封装断言函数
    def assertResponseCode(self,code,cell):
        excelcode=self.excelCode=self.readSheetdata(cell)
        try:
            assert excelcode==code
        except AssertionError as msg:
            print "断言失败"
            return False
        else:
            return True


    def getLoginUserMess(self):
        method="post"
        url="http://apptest.buddyniu.com/api/apps/BUDDY_API_TEST/accounts/login"
        data={"password":"e10adc3949ba59abbe56e057f20f883e","clientSecret":"a123af4e331cf61c0324cd43cbc2135d","accountId":"13590404631"}
        data = json.dumps(data)
        res=self.httpPost(data,url)
        return res

    def getAccesssToken(self):
        res=self.getLoginUserMess()
        return res['data']['accessToken']

    def getUserId(self):
        res=self.getLoginUserMess()
        return res['data']['userId']

    def getRefreshToken(self):
        res=self.getLoginUserMess()
        return res['data']['refreshToken']

    def getAfterUrl(self):
        userMess=self.getLoginUserMess()
        afterUrl="?access_token ="+userMess['data']['accessToken']+"&userId ="+userMess['data']['userId']
        return afterUrl




testClass=testClass()
keyword = {"phone":"13049008878"}
aaa=testClass.httpGet("https://sandbox.bitkinetic.com/api/v1/Register/testGetverifyCode",keyword)



